"""
Parse Excel/CSV files to extract test cases for import.
Supports .xlsx, .xls, and .csv with flexible column name matching.

SMART HEADER DETECTION for Excel:
- Single-row: Row 1 = headers, data from row 2
- Multi-row: Row 1 = category headers (skip), Row 2 = field names, data from row 3
- Auto-detect by counting field matches; if ambiguous, raise HeaderDetectionAmbiguous
- CSV: Row 1 = headers, data from row 2 (no detection needed)
"""
from __future__ import annotations

import csv
import io
import re
import logging
from typing import Any, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# Column name variations -> standard field key (case-insensitive)
COLUMN_MAPPINGS: List[Tuple[List[str], str]] = [
    (["test id", "id", "testid"], "test_id"),
    (["test scenario", "scenario"], "scenario"),
    (["test description", "description"], "description"),
    (["pre-condition", "precondition", "pre condition", "pre-test conditions"], "preconditions"),
    (["test data", "testdata", "data"], "test_data"),
    (["test steps", "test step", "step", "steps"], "steps"),
    (["expected result", "expected", "expected results"], "expected_result"),
    (["actual result", "actual", "actual results"], "actual_result"),
    (["status"], "status"),
]

PREFERRED_SHEET_NAMES = ("test cases", "testcases", "test case")

# Minimum field matches for valid header detection (Test ID, Test Steps, Expected Result, etc.)
MIN_HEADER_MATCHES = 3


class HeaderDetectionAmbiguous(Exception):
    """Raised when Excel header format cannot be determined. Carries row previews for user selection."""
    def __init__(self, row1_preview: List[str], row2_preview: List[str], row3_preview: List[str]):
        self.row1_preview = row1_preview
        self.row2_preview = row2_preview
        self.row3_preview = row3_preview
        super().__init__("Header format ambiguous. Please select header row.")


def _normalize(val: Any) -> str:
    """Trim whitespace and coerce to string."""
    if val is None:
        return ""
    return str(val).strip()


def _parse_steps(val: Any) -> List[str]:
    """Parse steps from string (pipe, newline) or list into list of step strings."""
    if val is None:
        return []
    if isinstance(val, list):
        return [_normalize(s) for s in val if s]
    s = _normalize(val)
    if not s or s.lower() in ("n/a", "none", "-"):
        return []
    parts = re.split(r"\s*\|\s*|\n+", s)
    result = []
    for p in parts:
        p = re.sub(r"^\d+[.)]\s*", "", p.strip())
        if p:
            result.append(p)
    return result if result else [s]


def _map_column(header: str) -> Optional[str]:
    """Map a header string to field key, or None if no match."""
    h = header.lower().strip()
    for variations, field_key in COLUMN_MAPPINGS:
        for v in variations:
            if v == h or v in h or h in v:
                return field_key
    return None


def _count_field_matches(row_values: List[str]) -> int:
    """Count how many expected field variations match in a row. Used for header detection."""
    seen_fields: set[str] = set()
    for val in row_values:
        h = _normalize(val).lower()
        if not h:
            continue
        for variations, field_key in COLUMN_MAPPINGS:
            if field_key in seen_fields:
                continue
            for v in variations:
                if v == h or v in h or h in v:
                    seen_fields.add(field_key)
                    break
    return len(seen_fields)


def _detect_excel_header_format(sheet: Worksheet) -> Tuple[Optional[int], Optional[int], str]:
    """
    Smart header detection for Excel.
    Decision: Row 1 high & >= Row 2 -> single-row. Row 2 high -> multi-row. Else unknown.
    Returns (header_row_1based, data_start_row_1based, format_str).
    """
    if sheet.max_row < 2:
        return None, None, "unknown"

    def row_values(r: int) -> List[str]:
        row = sheet[r]
        return [_normalize(c.value) for c in row if c.value]

    row1_vals = row_values(1)
    row2_vals = row_values(2)
    score1 = _count_field_matches(row1_vals)
    score2 = _count_field_matches(row2_vals)

    # Decision matrix
    if score1 >= MIN_HEADER_MATCHES and score1 >= score2:
        return 1, 2, "single-row"
    if score2 >= MIN_HEADER_MATCHES and score2 > score1:
        return 2, 3, "multi-row"
    if score1 >= MIN_HEADER_MATCHES and score2 >= MIN_HEADER_MATCHES and score1 == score2:
        return 1, 2, "single-row"  # Prefer simpler
    return None, None, "unknown"


def _get_excel_sheet(workbook) -> Optional[Worksheet]:
    """Get worksheet to use. Prefer 'Test Cases' sheet, else first sheet."""
    for sheet in workbook.worksheets:
        if sheet.title.lower() in PREFERRED_SHEET_NAMES and sheet.max_row >= 2:
            return sheet
    for sheet in workbook.worksheets:
        if sheet.max_row >= 2:
            return sheet
    return None


def _build_column_map(header_row) -> dict:
    """Build col_idx -> field_key from header row (0-based indices)."""
    col_map: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        val = _normalize(cell.value)
        if not val:
            continue
        field = _map_column(val)
        if field:
            col_map[idx] = field
    return col_map


def _row_to_test_case(row: List[Any], col_map: dict) -> Optional[dict]:
    """Convert data row to test case dict. Returns None if row is empty/invalid."""
    data: dict[str, Any] = {}
    for col_idx, field_key in col_map.items():
        if col_idx < len(row):
            val = row[col_idx]
            if val is not None and _normalize(val):
                data[field_key] = val

    scenario = _normalize(data.get("scenario") or data.get("test_id") or "")
    description = _normalize(data.get("description") or "")
    expected = _normalize(data.get("expected_result") or "")
    if not scenario and not description and not expected:
        return None

    steps = _parse_steps(data.get("steps"))
    if not steps:
        steps = ["Execute test scenario as described"]

    return {
        "test_id": _normalize(data.get("test_id") or "") or None,
        "scenario": scenario or description or "Test case",
        "description": description or scenario,
        "preconditions": _normalize(data.get("preconditions") or "") or "None",
        "test_data": _normalize(data.get("test_data") or "") or None,
        "steps": steps,
        "expected_result": expected or "Verify expected behavior",
        "actual_result": _normalize(data.get("actual_result") or "") or None,
        "status": _normalize(data.get("status") or "") or None,
    }


def _validate_and_normalize(cases: List[dict]) -> Tuple[List[dict], List[str]]:
    """Validate test cases, assign test_id if missing, return (normalized_cases, warnings)."""
    warnings: List[str] = []
    seen_ids: set[str] = set()
    result: List[dict] = []

    for i, tc in enumerate(cases):
        tid = tc.get("test_id") or ""
        if not tid:
            tid = f"TC_{i + 1:04d}"
            tc["test_id"] = tid
            warnings.append(f"Row {i + 1}: Missing Test ID, assigned '{tid}'")
        if tid in seen_ids:
            warnings.append(f"Row {i + 1}: Duplicate Test ID '{tid}'")
        seen_ids.add(tid)

        # Ensure steps is a list
        if isinstance(tc.get("steps"), str):
            tc["steps"] = _parse_steps(tc["steps"])
        if not tc.get("steps"):
            tc["steps"] = ["Execute test scenario as described"]

        result.append(tc)

    return result, warnings


def _get_row_preview(sheet: Worksheet, row_idx: int) -> List[str]:
    """Get cell values from a row as strings for preview."""
    row = sheet[row_idx]
    return [_normalize(c.value) for c in row]


def parse_excel(
    content: bytes,
    filename: str = "",
    header_row_override: Optional[int] = None,
) -> Tuple[List[dict], dict, List[str]]:
    """
    Parse Excel (.xlsx or .xls) file.
    header_row_override: 1 or 2 to force header row; None = auto-detect.
    """
    fn_lower = filename.lower()
    metadata: dict = {}
    warnings: List[str] = []

    if fn_lower.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
        sheet = _get_excel_sheet(wb)
        if sheet is None:
            sheet_names = [s.title for s in wb.worksheets]
            raise ValueError(
                f"No usable sheet found. Sheets: {sheet_names}. "
                "Excel files must have at least 2 rows."
            )

        if header_row_override is not None:
            # User manually selected header row
            h_row = int(header_row_override)
            if h_row not in (1, 2):
                raise ValueError("header_row must be 1 or 2")
            data_start = h_row + 1
            fmt = "user-selected"
        else:
            # Auto-detect
            h_row, data_start, fmt = _detect_excel_header_format(sheet)
            if fmt == "unknown":
                row1 = _get_row_preview(sheet, 1)
                row2 = _get_row_preview(sheet, 2)
                row3 = _get_row_preview(sheet, 3) if sheet.max_row >= 3 else []
                raise HeaderDetectionAmbiguous(row1, row2, row3)

        header_row = sheet[h_row]
        col_map = _build_column_map(header_row)

        if "scenario" not in col_map.values() and "expected_result" not in col_map.values():
            raise ValueError(
                f"Could not find required columns (Scenario or Expected Result) in row {h_row}."
            )

        metadata["sheet_used"] = sheet.title
        metadata["column_map"] = {str(k): v for k, v in col_map.items()}
        metadata["header_row"] = h_row
        metadata["data_start_row"] = data_start
        metadata["header_format"] = fmt

        cases: List[dict] = []
        for row_idx in range(data_start, sheet.max_row + 1):
            row_cells = sheet[row_idx]
            row = [c.value for c in row_cells]
            if all(v is None or not _normalize(v) for v in row):
                continue
            tc = _row_to_test_case(row, col_map)
            if tc:
                cases.append(tc)

        normalized, extra_warnings = _validate_and_normalize(cases)
        warnings.extend(extra_warnings)

        if not normalized:
            raise ValueError(f"No valid test case rows found in sheet '{sheet.title}'")

        return normalized, metadata, warnings

    elif fn_lower.endswith(".xls"):
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=content)
            sheet = wb.sheet_by_index(0)
            rows: List[List[Any]] = []
            for r in range(sheet.nrows):
                row = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                rows.append(row)
            # .xls: use detection or override
            if header_row_override is not None:
                h_idx = int(header_row_override) - 1  # 1-based -> 0-based
                if h_idx not in (0, 1):
                    raise ValueError("header_row must be 1 or 2")
                return _parse_excel_rows(rows, metadata, warnings, filename, header_row_index=h_idx)
            # Auto-detect for .xls
            row1_vals = [_normalize(v) for v in (rows[0] if rows else [])]
            row2_vals = [_normalize(v) for v in (rows[1] if len(rows) >= 2 else [])]
            s1 = _count_field_matches(row1_vals)
            s2 = _count_field_matches(row2_vals)
            if s1 >= MIN_HEADER_MATCHES and s1 >= s2:
                return _parse_excel_rows(rows, metadata, warnings, filename, header_row_index=0)
            if s2 >= MIN_HEADER_MATCHES and s2 > s1:
                return _parse_excel_rows(rows, metadata, warnings, filename, header_row_index=1)
            if s1 >= MIN_HEADER_MATCHES and s2 >= MIN_HEADER_MATCHES:
                return _parse_excel_rows(rows, metadata, warnings, filename, header_row_index=0)
            row3 = [_normalize(v) for v in (rows[2] if len(rows) >= 3 else [])]
            raise HeaderDetectionAmbiguous(row1_vals, row2_vals, row3)
        except ImportError:
            raise ValueError("Reading .xls requires 'xlrd'. Install with: pip install xlrd")
    else:
        raise ValueError("Unsupported Excel format. Use .xlsx or .xls")


def _parse_excel_rows(
    rows: List[List[Any]],
    metadata: dict,
    warnings: List[str],
    filename: str,
    header_row_index: int = 1,
) -> Tuple[List[dict], dict, List[str]]:
    """Parse Excel row list (.xls). header_row_index: 0 = row 1, 1 = row 2."""
    if len(rows) < header_row_index + 1:
        raise ValueError("Excel file does not have enough rows")
    header_idx = header_row_index
    header_row = rows[header_idx]
    col_map: dict[int, str] = {}
    for c, val in enumerate(header_row):
        h = _normalize(val)
        if not h:
            continue
        f = _map_column(h)
        if f:
            col_map[c] = f
    if "scenario" not in col_map.values() and "expected_result" not in col_map.values():
        raise ValueError(
            "Could not find required columns (Scenario or Expected Result)."
        )
    cases: List[dict] = []
    for row in rows[header_idx + 1:]:  # Data from row after header
        tc = _row_to_test_case(row, col_map)
        if tc:
            cases.append(tc)
    normalized, extra_warnings = _validate_and_normalize(cases)
    warnings.extend(extra_warnings)
    if not normalized:
        raise ValueError("No valid test case rows found (data starts at row 3)")
    return normalized, metadata, warnings


def _parse_csv_rows(
    rows: List[List[Any]], metadata: dict, warnings: List[str], filename: str
) -> Tuple[List[dict], dict, List[str]]:
    """
    Parse CSV row list: Row 1 (index 0) = headers, Row 2+ (index 1+) = data.
    No row skipping for CSV.
    """
    if not rows:
        raise ValueError("File is empty")
    header_row = rows[0]
    col_map: dict[int, str] = {}
    for c, val in enumerate(header_row):
        h = _normalize(val)
        if not h:
            continue
        f = _map_column(h)
        if f:
            col_map[c] = f
    if "scenario" not in col_map.values() and "expected_result" not in col_map.values():
        raise ValueError("No test case headers found in row 1. Expected: Test ID, Scenario, Expected Result")
    cases: List[dict] = []
    for row in rows[1:]:  # Data from row 2 onwards
        tc = _row_to_test_case(row, col_map)
        if tc:
            cases.append(tc)
    normalized, extra_warnings = _validate_and_normalize(cases)
    warnings.extend(extra_warnings)
    if not normalized:
        raise ValueError("No valid test case rows found")
    metadata["header_format"] = "csv-standard"
    return normalized, metadata, warnings


def parse_csv(content: bytes, filename: str = "") -> Tuple[List[dict], dict, List[str]]:
    """Parse CSV file. Returns (test_cases, metadata, warnings)."""
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    # CSV: Row 1 = headers, data from row 2 (no row skipping)
    return _parse_csv_rows(rows, {}, [], filename)


def parse_file(
    content: bytes,
    filename: str,
    header_row_override: Optional[int] = None,
) -> Tuple[List[dict], dict, List[str]]:
    """
    Auto-detect format and parse. Returns (test_cases, metadata, warnings).
    header_row_override: 1 or 2 for Excel; None = auto-detect.
    """
    fn = filename.lower()
    if fn.endswith(".csv"):
        return parse_csv(content, filename)
    if fn.endswith(".xlsx") or fn.endswith(".xls"):
        return parse_excel(content, filename, header_row_override)
    raise ValueError("File format not supported. Use .xlsx, .xls, or .csv")
