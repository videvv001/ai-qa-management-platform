"""
Merge test cases into an uploaded Excel template.
Matches exact template structure: Summary sheet unchanged, Test Cases sheet
has rows 1-2 as headers (with merges), row 3+ as data. Columns A-L (1-12).
"""
from __future__ import annotations

import re
import tempfile
from copy import copy
from pathlib import Path
from typing import Any, List, Mapping, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

# Template: columns A-L (1-12); row 1-2 headers, row 3+ data
MAX_DATA_COLS = 12
SHEET_NAME = "Test Cases"
SUMMARY_SHEET = "Summary"
MAX_TEMPLATE_SIZE_BYTES = 10 * 1024 * 1024  # 10MB


def format_test_steps(steps_input: str | List[str] | None) -> str:
    """
    Format test steps as enumerated list.
    Input: "Navigate to page | Click button | Verify result"
    Output: "1. Navigate to page\n2. Click button\n3. Verify result"
    Handles N/A, None; strips existing numbering.
    """
    if steps_input is None:
        return ""
    if isinstance(steps_input, list):
        parts = [str(s).strip() for s in steps_input if s]
    else:
        s = str(steps_input).strip()
        if not s or s in ("N/A", "None", ""):
            return ""
        parts = [p.strip() for p in s.split("|") if p.strip()]
    if not parts:
        return ""
    formatted = []
    for i, step in enumerate(parts, 1):
        step = step.strip()
        step = re.sub(r"^\d+[.)]\s*", "", step)  # strip existing numbering
        formatted.append(f"{i}. {step}")
    return "\n".join(formatted)


def _feature_prefix(feature_name: str) -> str:
    """
    Generate prefix for Test IDs based on feature name.
    Examples: "login page" -> "LOGIN", "helper management" -> "HM"
    """
    if not feature_name or not feature_name.strip():
        return "GEN"
    stop_words = {"page", "the", "a", "an"}
    words = [w for w in feature_name.lower().split() if w not in stop_words]
    if not words:
        return feature_name[:5].upper() if len(feature_name) >= 5 else feature_name.upper()
    if len(words) == 1:
        return words[0][:5].upper()
    return "".join(w[0].upper() for w in words[:2])


def _get_style_dict(ws: Worksheet, template_row: int) -> dict[int, dict[str, Any]]:
    """Copy font, fill, border, alignment, number_format from template row (cols 1-12)."""
    styles: dict[int, dict[str, Any]] = {}
    for col in range(1, MAX_DATA_COLS + 1):
        cell = ws.cell(row=template_row, column=col)
        styles[col] = {}
        if cell.has_style:
            styles[col] = {
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
            }
        styles[col]["number_format"] = cell.number_format
    return styles


def _apply_style(cell, style: dict[str, Any]) -> None:
    if not style:
        return
    if "font" in style:
        cell.font = style["font"]
    if "fill" in style:
        cell.fill = style["fill"]
    if "border" in style:
        cell.border = style["border"]
    if "alignment" in style:
        cell.alignment = style["alignment"]
    if "number_format" in style:
        cell.number_format = style["number_format"]


def _copy_row(
    ws_src: Worksheet,
    row_src: int,
    ws_dest: Worksheet,
    row_dest: int,
    *,
    include_number_format: bool = True,
) -> None:
    """Copy one row (value + style) from source to destination sheet (cols 1-12)."""
    for col in range(1, MAX_DATA_COLS + 1):
        src_cell = ws_src.cell(row=row_src, column=col)
        dest_cell = ws_dest.cell(row=row_dest, column=col)
        dest_cell.value = src_cell.value
        if src_cell.has_style:
            style: dict[str, Any] = {
                "font": copy(src_cell.font),
                "fill": copy(src_cell.fill),
                "border": copy(src_cell.border),
                "alignment": copy(src_cell.alignment),
            }
            if include_number_format:
                style["number_format"] = src_cell.number_format
            _apply_style(dest_cell, style)


def _tc_value(tc: Mapping[str, Any], key: str, default: str = "") -> str:
    """Get string from dict with snake_case or camelCase key."""
    v = tc.get(key) or tc.get(key.replace("_", ""))
    if v is None and "_" in key:
        c = "".join(w.capitalize() if i else w.lower() for i, w in enumerate(key.split("_")))
        v = tc.get(c)
    return str(v).strip() if v is not None else default


def _write_data_row(
    ws: Worksheet,
    row: int,
    tc: Mapping[str, Any],
    feature_prefix: str,
    idx: int,
    template_styles: dict[int, dict[str, Any]],
    *,
    global_no: int | None = None,
) -> None:
    """Write one test case row (row 3+), columns A-L, and apply template formatting.
    idx = number within feature (for Test ID). global_no = optional sequential No. across all (column A).
    """
    scenario = _tc_value(tc, "test_scenario") or _tc_value(tc, "testScenario")
    description = _tc_value(tc, "test_description") or _tc_value(tc, "description")
    precondition = _tc_value(tc, "pre_condition") or _tc_value(tc, "precondition")
    test_data = _tc_value(tc, "test_data") or _tc_value(tc, "testData")
    steps_raw = tc.get("test_steps") or tc.get("testSteps")
    if isinstance(steps_raw, list):
        steps_str = format_test_steps(steps_raw)
    else:
        steps_str = format_test_steps(steps_raw if isinstance(steps_raw, str) else "")
    expected = _tc_value(tc, "expected_result") or _tc_value(tc, "expectedResult")

    test_id = f"TC_{feature_prefix}_{str(idx).zfill(3)}"
    no_value = global_no if global_no is not None else idx

    ws.cell(row=row, column=1, value=no_value)
    ws.cell(row=row, column=2, value=test_id)
    ws.cell(row=row, column=3, value=scenario)
    ws.cell(row=row, column=4, value=description)
    ws.cell(row=row, column=5, value=precondition)
    ws.cell(row=row, column=6, value=test_data)
    ws.cell(row=row, column=7, value=steps_str)
    ws.cell(row=row, column=8, value=expected)
    ws.cell(row=row, column=9, value="")
    ws.cell(row=row, column=10, value="Not Executed")
    ws.cell(row=row, column=11, value="")
    ws.cell(row=row, column=12, value="")

    for col in range(1, MAX_DATA_COLS + 1):
        _apply_style(ws.cell(row=row, column=col), template_styles.get(col, {}))


def merge_test_cases_to_excel(
    template_path: str | Path,
    test_cases: List[Mapping[str, Any]],
    feature_name: str,
) -> str:
    """
    Export test cases for a single feature to the template.
    - Summary sheet is left unchanged.
    - Test Cases sheet: rows 1-2 kept as-is, replace data from row 3+ with new test cases.
    - Preserve all formatting (font, fill, border, alignment, number_format).
    - Columns A-L: No., Test ID, Test Scenario, Test Description, Pre-condition, Test Data,
      Step (enumerated), Expected Result, Actual Result, Status, Comment, (empty).
    Returns path to saved temporary .xlsx file.
    """
    path = Path(template_path)
    if not path.exists() or path.stat().st_size > MAX_TEMPLATE_SIZE_BYTES:
        raise ValueError("Template file missing or exceeds size limit")

    wb = load_workbook(path, read_only=False)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Template must contain a sheet named '{SHEET_NAME}'")
    ws = wb[SHEET_NAME]

    header_end_row = 2
    data_start_row = 3
    template_styles = _get_style_dict(ws, data_start_row) if ws.max_row >= data_start_row else {}

    # Delete existing data rows only (keep headers rows 1-2)
    if ws.max_row >= data_start_row:
        ws.delete_rows(data_start_row, ws.max_row - header_end_row)

    feature_prefix = _feature_prefix(feature_name or "Export")

    for idx, tc in enumerate(test_cases, 1):
        row = data_start_row + idx - 1
        _write_data_row(ws, row, tc, feature_prefix, idx, template_styles)

    tmp = tempfile.NamedTemporaryFile(
        prefix="export_test_cases_",
        suffix=".xlsx",
        delete=False,
    )
    out_path = Path(tmp.name)
    tmp.close()
    wb.save(str(out_path))
    return str(out_path)


def merge_all_features_to_excel(
    template_path: str | Path,
    features_data: List[Tuple[str, List[Mapping[str, Any]]]],
) -> str:
    """
    Export all features into the template's single "Test Cases" sheet.
    - Summary sheet is kept from template (unchanged).
    - Test Cases sheet is kept; only data rows (row 3+) are replaced.
    - All features' test cases are combined in order: Feature1's cases first, then Feature2's, etc.
    - Column A (No.) = global sequential number (1, 2, 3, ...).
    - Column B (Test ID) = per-feature (e.g. TC_FEAT1_001, ..., TC_FEAT2_001, ...).
    Returns path to saved temporary .xlsx file.
    """
    path = Path(template_path)
    if not path.exists() or path.stat().st_size > MAX_TEMPLATE_SIZE_BYTES:
        raise ValueError("Template file missing or exceeds size limit")

    wb = load_workbook(path, read_only=False)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Template must contain a sheet named '{SHEET_NAME}'")
    ws = wb[SHEET_NAME]

    header_end_row = 2
    data_start_row = 3
    template_styles = _get_style_dict(ws, data_start_row) if ws.max_row >= data_start_row else {}

    # Delete existing data rows only (keep headers rows 1-2)
    if ws.max_row >= data_start_row:
        ws.delete_rows(data_start_row, ws.max_row - header_end_row)

    global_row = 1  # Sequential No. across all features (1, 2, 3, ...)
    current_data_row = data_start_row  # Next row to write (3, 4, 5, ...)

    for feature_name, test_cases in features_data:
        feature_prefix = _feature_prefix(feature_name or "Feature")
        for idx, tc in enumerate(test_cases, 1):
            _write_data_row(
                ws,
                current_data_row,
                tc,
                feature_prefix,
                idx,
                template_styles,
                global_no=global_row,
            )
            global_row += 1
            current_data_row += 1

    tmp = tempfile.NamedTemporaryFile(
        prefix="export_all_test_cases_",
        suffix=".xlsx",
        delete=False,
    )
    out_path = Path(tmp.name)
    tmp.close()
    wb.save(str(out_path))
    return str(out_path)


# --- Module-based export with auto-detection ---

# Header (row 2) text patterns -> field key for column mapping
# Order matters for matching: more specific first
_HEADER_PATTERNS: List[Tuple[List[str], str]] = [
    (["no", "no."], "no"),
    (["test id"], "test_id"),
    (["test scenario", "scenario"], "scenario"),
    (["test description", "description"], "description"),
    (["pre-condition", "precondition"], "preconditions"),
    (["test data", "data"], "test_data"),
    (["step", "test step"], "steps"),
    (["expected result", "expected"], "expected_result"),
    (["actual result", "actual"], "actual_result"),
    (["status"], "status"),
    (["comment", "notes"], "notes"),
]


def _detect_column_mapping(ws: Worksheet, header_row: int) -> dict[int, str]:
    """
    Scan header_row for column headers (case-insensitive) and return
    col_idx -> field_key mapping. Field keys: no, test_id, scenario, etc.
    """
    mapping: dict[int, str] = {}
    max_col = min(ws.max_column, 50) if ws.max_column else 50
    for col in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col)
        val = (cell.value or "").strip()
        if not val:
            continue
        val_lower = val.lower()
        for patterns, field_key in _HEADER_PATTERNS:
            if val_lower in patterns:
                mapping[col] = field_key
                break
    return mapping


def _find_header_location(
    wb: Any,
    prefer_multi_row: bool = True,
) -> Tuple[Worksheet, int, int, dict[int, str]]:
    """
    Scan workbook for target field headers.
    Process: first tab -> row 1, row 2; if no target fields, next tab -> row 1, row 2; etc.
    prefer_multi_row=True: try row 2 first (multi-row), then row 1 (single-row) per sheet
    prefer_multi_row=False: try row 1 first (single-row), then row 2 (multi-row) per sheet
    Returns (worksheet, header_row, data_start_row, col_map).
    """
    sheet_names = list(wb.sheetnames)
    if SHEET_NAME in sheet_names and sheet_names[0] != SHEET_NAME:
        sheet_names = [SHEET_NAME] + [s for s in sheet_names if s != SHEET_NAME]
    row_order = (2, 1) if prefer_multi_row else (1, 2)
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for header_row in row_order:
            col_map = _detect_column_mapping(ws, header_row)
            if len(col_map) >= 2:
                data_start_row = header_row + 1
                return (ws, header_row, data_start_row, col_map)

    raise ValueError(
        "Could not find target field headers in any sheet. "
        "Expected headers like: Test ID, Test Scenario, Expected Result, etc. "
        "Check row 1 and row 2 of each sheet."
    )


def _do_merge_into_template(
    wb: Any,
    test_cases: List[Mapping[str, Any]],
    ws: "Worksheet",
    header_end_row: int,
    data_start_row: int,
    col_map: dict[int, str],
) -> str:
    """
    Write test cases into worksheet. Caller provides workbook, sheet, header location.
    """
    template_styles = _get_style_dict(ws, data_start_row) if ws.max_row >= data_start_row else {}

    if ws.max_row >= data_start_row:
        ws.delete_rows(data_start_row, ws.max_row - header_end_row)

    for idx, tc in enumerate(test_cases, 1):
        row = data_start_row + idx - 1
        latest = tc.get("latest_execution") or {}
        steps_raw = tc.get("steps") or []
        steps_str = format_test_steps(steps_raw) if isinstance(steps_raw, list) else format_test_steps(steps_raw)

        values: dict[str, Any] = {
            "no": idx,
            "test_id": str(tc.get("test_id") or ""),
            "scenario": str(tc.get("scenario") or ""),
            "description": str(tc.get("description") or ""),
            "preconditions": str(tc.get("preconditions") or ""),
            "test_data": str(tc.get("test_data") or ""),
            "steps": steps_str,
            "expected_result": str(tc.get("expected_result") or ""),
            "actual_result": str(latest.get("actual_result") or ""),
            "status": str(latest.get("status") or "Not Executed"),
            "notes": str(latest.get("notes") or ""),
        }

        for col, field_key in col_map.items():
            cell = ws.cell(row=row, column=col)
            cell.value = values.get(field_key, "")
            _apply_style(cell, template_styles.get(col, {}))

    tmp = tempfile.NamedTemporaryFile(
        prefix="module_export_",
        suffix=".xlsx",
        delete=False,
    )
    out_path = Path(tmp.name)
    tmp.close()
    wb.save(str(out_path))
    return str(out_path)


def merge_module_cases_to_excel_template(
    template_path: str | Path,
    test_cases: List[Mapping[str, Any]],
    module_name: str,
    *,
    prefer_multi_row: bool = True,
) -> str:
    """
    Merge module test cases into an uploaded Excel template.
    Auto-detects target fields: scans each sheet, checks row 1 then row 2.
    prefer_multi_row: try multi-row headers (row 2) first, else single-row (row 1).
    """
    path = Path(template_path)
    if not path.exists() or path.stat().st_size > MAX_TEMPLATE_SIZE_BYTES:
        raise ValueError("Template file missing or exceeds size limit")

    wb = load_workbook(path, read_only=False)
    ws, header_row, data_start_row, col_map = _find_header_location(wb, prefer_multi_row=prefer_multi_row)
    return _do_merge_into_template(wb, test_cases, ws, header_row, data_start_row, col_map)


def merge_module_cases_to_excel_template_with_fallback(
    template_path: str | Path,
    test_cases: List[Mapping[str, Any]],
    module_name: str,
) -> str:
    """
    Merge module test cases into template. Tries multi-row first, falls back to single-row on error.
    """
    multi_row_error = None
    try:
        return merge_module_cases_to_excel_template(
            template_path, test_cases, module_name, prefer_multi_row=True
        )
    except Exception as e:
        multi_row_error = e
        try:
            return merge_module_cases_to_excel_template(
                template_path, test_cases, module_name, prefer_multi_row=False
            )
        except Exception as single_row_error:
            raise ValueError(
                f"Merge failed with both formats. "
                f"Multi-row: {multi_row_error}. Single-row: {single_row_error}"
            ) from single_row_error


# --- Template-less Excel export (single-row or multi-row header) ---

# Column headers for export (9 columns)
EXPORT_FIELD_HEADERS = [
    "Test ID",
    "Test Scenario",
    "Test Description",
    "Pre-condition",
    "Test Data",
    "Step",
    "Expected Result",
    "Actual Result",
    "Status",
]

# Multi-row: category row (row 1) merge ranges and labels
# Cols 1-2: Test Identification, 3-4: Test Description, 5: Pre-Test Conditions, 6: Test Data, 7-9: Test Execution
_CATEGORY_MERGES = [(1, 2), (3, 4), (5, 5), (6, 6), (7, 9)]
_CATEGORY_LABELS = ["Test Identification", "Test Description", "Pre-Test Conditions", "Test Data", "Test Execution"]

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue
_CATEGORY_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark blue
_HEADER_FONT = Font(bold=True, size=11)
_CATEGORY_FONT = Font(bold=True, size=12)


def _tc_to_row_values(tc: Mapping[str, Any]) -> List[Any]:
    """Convert one test case dict to a row of 9 values for export columns."""
    latest = tc.get("latest_execution") or {}
    steps_raw = tc.get("steps") or []
    steps_str = format_test_steps(steps_raw) if isinstance(steps_raw, list) else format_test_steps(steps_raw)
    return [
        str(tc.get("test_id") or ""),
        str(tc.get("scenario") or ""),
        str(tc.get("description") or ""),
        str(tc.get("preconditions") or ""),
        str(tc.get("test_data") or ""),
        steps_str,
        str(tc.get("expected_result") or ""),
        str(latest.get("actual_result") or ""),
        str(latest.get("status") or "Not Executed"),
    ]


def _apply_borders(ws: Worksheet, max_row: int, max_col: int = 9) -> None:
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = _THIN_BORDER


def build_module_cases_excel(
    test_cases: List[Mapping[str, Any]],
    module_name: str,
    header_format: str = "multi-row",
) -> str:
    """
    Build an Excel file from test cases without a template.
    header_format: 'single-row' (row 1 = headers, row 2+ = data) or 'multi-row' (row 1 = categories, row 2 = headers, row 3+ = data).
    Returns path to saved temporary .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    col_widths = [15, 25, 30, 25, 20, 30, 30, 30, 15]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w

    if header_format == "single-row":
        # Row 1: field headers only
        for col, label in enumerate(EXPORT_FIELD_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="middle")
        ws.row_dimensions[1].height = 20
        data_start = 2
    else:
        # Multi-row: Row 1 = category headers (merged), Row 2 = field headers
        for i, (merge_start, merge_end) in enumerate(_CATEGORY_MERGES):
            if merge_start < merge_end:
                ws.merge_cells(
                    start_row=1, start_column=merge_start,
                    end_row=1, end_column=merge_end,
                )
            cell = ws.cell(row=1, column=merge_start)
            cell.value = _CATEGORY_LABELS[i]
            cell.font = _CATEGORY_FONT
            cell.fill = _CATEGORY_FILL
            cell.alignment = Alignment(horizontal="center", vertical="middle")
        ws.row_dimensions[1].height = 25
        for col, label in enumerate(EXPORT_FIELD_HEADERS, 1):
            cell = ws.cell(row=2, column=col, value=label)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="middle")
        ws.row_dimensions[2].height = 20
        data_start = 3

    for idx, tc in enumerate(test_cases, 1):
        row_num = data_start + idx - 1
        for col, val in enumerate(_tc_to_row_values(tc), 1):
            ws.cell(row=row_num, column=col, value=val)

    _apply_borders(ws, data_start + len(test_cases) - 1 if test_cases else data_start)

    tmp = tempfile.NamedTemporaryFile(
        prefix="export_excel_",
        suffix=".xlsx",
        delete=False,
    )
    out_path = Path(tmp.name)
    tmp.close()
    wb.save(str(out_path))
    return str(out_path)


def build_combined_modules_excel(
    cases_by_module: List[Tuple[str, List[Mapping[str, Any]]]],
    header_format: str = "multi-row",
) -> str:
    """
    Build one Excel file with test cases from multiple modules.
    cases_by_module: list of (module_name, test_cases).
    Optionally include a "Module" column when len(cases_by_module) > 1.
    Returns path to saved temporary .xlsx file.
    """
    all_cases: List[Tuple[str, Mapping[str, Any]]] = []  # (module_name, tc)
    for mod_name, cases in cases_by_module:
        for tc in cases:
            all_cases.append((mod_name, tc))

    # Use first module name for "sheet" / naming; we don't add Module column for now to keep 9 columns
    first_name = cases_by_module[0][0] if cases_by_module else "Combined"
    # Build as single flat list (no module column per user spec); module grouping is just the order
    flat_cases = [tc for _, tc in all_cases]
    return build_module_cases_excel(flat_cases, first_name, header_format)
